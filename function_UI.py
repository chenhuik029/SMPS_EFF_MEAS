import math
from PyQt5.QtWidgets import QMainWindow, QDialog
from PyQt5.QtCore import QThread, QObject, pyqtSignal
from PyQt5.QtGui import QDoubleValidator
from UIpy import Main_ui, Support, Measurement_Setup
from function_msgbox import msg_box_ok, msg_box_auto_close, msg_box_ok_cancel
from Instrument_PyVisa import Basic_PyVisa, PS_Kikusui_PyVisa, Eload_Chroma_PyVisa, DMM_Keysight_PyVisa
from tkinter import filedialog
import re
import os
import time
import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.chart import ScatterChart, Reference, Series
import tkinter as tk
import datetime
thread_running = False


# Main UI for Fixed VIN, Variable Vout Test
class FixedVIN_VarVOUT_UI(QMainWindow, Main_ui.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.default_configuration()
        self.list_instrument = Basic_PyVisa.Basic_PyVisa()
        self.combobox_equipment_list()

        self.error_count = 0
        self.PS_address = ""
        self.PS_target_Vin = 0
        self.PS_limit_Vin = 0
        self.PS_limit_Cin = 0
        self.ELoad_IMAX = 0
        self.ELoad_ISTEP = 0
        self.ELoad_ISTART = 0
        self.ELoad_address = ""
        self.Eload_channel = ""
        self.DMM_VIN_address = ""
        self.DMM_CIN_address = ""
        self.DMM_VOUT_address = ""
        self.DMM_COUT_address = ""
        self.thread_count = 0

        self.checkBox_ExtSupUsed.clicked.connect(self.Ext_Supply_used_checked)
        self.checkBox_ReadIntVol_PSEquip.clicked.connect(self.DMM1_used_checked)
        self.checkBox_ReadIntCur_PSEquip.clicked.connect(self.DMM2_used_checked)
        self.checkBox_ReadOutVol_ELoad.clicked.connect(self.DMM3_used_checked)
        self.checkBox_ReadOutCur_ELoad.clicked.connect(self.DMM4_used_checked)
        self.pushButton_StartTest.clicked.connect(self.start_test)
        self.pushButton_AbortTest.clicked.connect(self.stop_test)
        self.pushButton_Browse.clicked.connect(self.browse_directory)
        self.pushButton_Refresh_Instrument.clicked.connect(self.refresh_instruments)

        self.lineEdit_IMAX.textChanged.connect(self.check_eload_input_value)
        self.lineEdit_ISTART.textChanged.connect(self.check_eload_input_value)
        self.lineEdit_ISTEP.textChanged.connect(self.check_eload_input_value)

        self.lineEdit_PS_VIN.textChanged.connect(self.check_PS_input_value)
        self.lineEdit_PS_VIN_Limit.textChanged.connect(self.check_PS_input_value)

        self.actionSupported_Device.triggered.connect(self.supported_device)
        self.actionExit.triggered.connect(self.close_app)
        self.actionMeasurement_Setup.triggered.connect(self.measurement_setup)
        self.actionSupport.triggered.connect(self.support)

    # Exit the application
    def close_app(self):
        self.close()

    # Show Supported Device
    def supported_device(self):
        msg_box_ok("Supported Device:\n\n"
                   "Power Supply: \n"
                   "- Kikusui PBZ 20-20\n\n"
                   "Digital Multimeter:\n"
                   "- Keysight DMM\n\n"
                   "Electronic Load\n"
                   "- Chroma ELOAD")

    # Show measurement setup diagram
    def measurement_setup(self):
        meas_setup_ui = MEAS_SETUP_GUIDE()
        meas_setup_ui.exec_()

    # Support Contact
    def support(self):
        support_ui = ABOUT_UI()
        support_ui.exec_()

    # Default configuration setting
    def default_configuration(self):
        self.Ext_Supply_used_checked()
        self.onlyFloat = QDoubleValidator()
        self.lineEdit_PS_VIN.setValidator(self.onlyFloat)
        self.lineEdit_PS_VIN_Limit.setValidator(self.onlyFloat)
        self.lineEdit_PS_Cur_Limit.setValidator(self.onlyFloat)
        self.lineEdit_IMAX.setValidator(self.onlyFloat)
        self.lineEdit_ISTEP.setValidator(self.onlyFloat)
        self.lineEdit_ISTART.setValidator(self.onlyFloat)

    # Refresh Instruments
    def refresh_instruments(self):
        self.combobox_equipment_list()
        msg_box_auto_close("Searching equipment ... ...")

    # Check Input value of ELoad
    def check_eload_input_value(self):
        if not self.lineEdit_ISTART.text() == "" and not self.lineEdit_IMAX.text() == "":
            if float(self.lineEdit_ISTART.text()) > float(self.lineEdit_IMAX.text()):
                msg_box_ok("Start load current must be SMALLER than Max current!!")
                self.lineEdit_ISTART.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 10);}")
                return False
            else:
                self.lineEdit_ISTART.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                self.lineEdit_ISTEP.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                return True
        if not self.lineEdit_ISTEP.text() == "" and not self.lineEdit_IMAX.text() == "":
            if float(self.lineEdit_ISTEP.text()) > float(self.lineEdit_IMAX.text()):
                msg_box_ok("Step load current must be SMALLER than Max current!!")
                self.lineEdit_ISTEP.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 10);}")
                return False
            else:
                self.lineEdit_ISTART.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                self.lineEdit_ISTEP.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                return True

    # Check Input value of ELoad
    def check_PS_input_value(self):
        if not self.lineEdit_PS_VIN_Limit.text() == "" and not self.lineEdit_PS_VIN.text() == "":
            if float(self.lineEdit_PS_VIN.text()) > float(self.lineEdit_PS_VIN_Limit.text()):
                msg_box_ok("Suupply Voltage must be SMALLER than Supply Voltage Limit!!")
                self.lineEdit_PS_VIN.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 10);}")
                return False
            else:
                self.lineEdit_PS_VIN.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                return True

    # Initial check box status
    def Ext_Supply_used_checked(self):
        if self.checkBox_ExtSupUsed.isChecked():
            self.comboBox_PS_Address.setEnabled(True)
            self.lineEdit_PS_VIN.setEnabled(True)
            self.lineEdit_PS_VIN_Limit.setEnabled(True)
            self.lineEdit_PS_Cur_Limit.setEnabled(True)
            self.checkBox_ReadIntCur_PSEquip.setEnabled(True)
            self.label_23.setEnabled(True)
            self.checkBox_ReadIntVol_PSEquip.setEnabled(True)
            self.label_19.setEnabled(True)
            if self.checkBox_ReadIntVol_PSEquip.isChecked():
                self.comboBox_DMM_VI.setDisabled(True)
            else:
                self.comboBox_DMM_VI.setDisabled(False)
            if self.checkBox_ReadIntCur_PSEquip.isChecked():
                self.comboBox_DMM_CI.setDisabled(True)
            else:
                self.comboBox_DMM_CI.setDisabled(False)
        else:
            self.comboBox_PS_Address.setEnabled(False)
            self.lineEdit_PS_VIN.setEnabled(False)
            self.lineEdit_PS_VIN_Limit.setEnabled(False)
            self.lineEdit_PS_Cur_Limit.setEnabled(False)
            self.checkBox_ReadIntCur_PSEquip.setEnabled(False)
            self.label_23.setEnabled(False)
            self.checkBox_ReadIntVol_PSEquip.setEnabled(False)
            self.label_19.setEnabled(False)
            self.comboBox_DMM_VI.setDisabled(False)
            self.comboBox_DMM_CI.setDisabled(False)

    # Initial check box status
    def DMM1_used_checked(self):
        if self.checkBox_ReadIntVol_PSEquip.isChecked():
            self.comboBox_DMM_VI.setEnabled(False)
        else:
            self.comboBox_DMM_VI.setEnabled(True)

    # Initial check box status
    def DMM2_used_checked(self):
        if self.checkBox_ReadIntCur_PSEquip.isChecked():
            self.comboBox_DMM_CI.setEnabled(False)
        else:
            self.comboBox_DMM_CI.setEnabled(True)

    # Initial check box status
    def DMM3_used_checked(self):
        if self.checkBox_ReadOutVol_ELoad.isChecked():
            self.comboBox_DMM_VO.setEnabled(False)
        else:
            self.comboBox_DMM_VO.setEnabled(True)

    # Initial check box status
    def DMM4_used_checked(self):
        if self.checkBox_ReadOutCur_ELoad.isChecked():
            self.comboBox_DMM_CO.setEnabled(False)
        else:
            self.comboBox_DMM_CO.setEnabled(True)

    # List the equipment list to drop box
    def combobox_equipment_list(self):
        self.comboBox_PS_Address.clear()
        self.comboBox_ELoad_Address.clear()
        self.comboBox_DMM_VI.clear()
        self.comboBox_DMM_CI.clear()
        self.comboBox_DMM_VO.clear()
        self.comboBox_DMM_CO.clear()
        self.scan_equipment_list()

        if self.comboBox_PS_Address.currentIndex() < 0:
            self.comboBox_PS_Address.addItem("Please select the targeted instrument")
            self.comboBox_PS_Address.setCurrentIndex(0)

        if self.comboBox_ELoad_Address.currentIndex() < 0:
            self.comboBox_ELoad_Address.addItem("Please select the targeted instrument")
            self.comboBox_ELoad_Address.setCurrentIndex(0)

        if self.comboBox_ELoad_CH.currentIndex() < 0:
            self.comboBox_ELoad_CH.addItems(["Select a channel", '1', '2', '3', '4'])
            self.comboBox_ELoad_CH.setCurrentIndex(0)

        if self.comboBox_DMM_VI.currentIndex() < 0:
            self.comboBox_DMM_VI.addItem("Please select the targeted instrument")
            self.comboBox_DMM_VI.setCurrentIndex(0)

        if self.comboBox_DMM_CI.currentIndex() < 0:
            self.comboBox_DMM_CI.addItem("Please select the targeted instrument")
            self.comboBox_DMM_CI.setCurrentIndex(0)

        if self.comboBox_DMM_VO.currentIndex() < 0:
            self.comboBox_DMM_VO.addItem("Please select the targeted instrument")
            self.comboBox_DMM_VO.setCurrentIndex(0)

        if self.comboBox_DMM_CO.currentIndex() < 0:
            self.comboBox_DMM_CO.addItem("Please select the targeted instrument")
            self.comboBox_DMM_CO.setCurrentIndex(0)

        for equipment in self.equipment_list:

            # To filter USB connection
            pattern_USB = re.compile("USB[0-9]")
            if pattern_USB.match(equipment.split("::", 4)[0]):
                # Filter Kikusui PBZ20-20 equipment for PS combobox
                if equipment.split("::", 4)[1] == "0x0B3E" and equipment.split("::", 4)[2] == "0x1012":
                    self.comboBox_PS_Address.addItem(equipment)
                # Add the remaining list to all the other DMM combobox
                else:
                    self.comboBox_DMM_VI.addItem(equipment)
                    self.comboBox_DMM_CI.addItem(equipment)
                    self.comboBox_DMM_VO.addItem(equipment)
                    self.comboBox_DMM_CO.addItem(equipment)

            # To filter non-USB connection
            pattern_ASRL = re.compile("ASRL[0-9]")
            if pattern_ASRL.match(equipment.split("::", 4)[0]):
                # Filter ELoad for ELoad combobox
                if equipment.split("::", 4)[1] == "INSTR":
                    self.comboBox_ELoad_Address.addItem(equipment)

    # Scan the equipment list
    def scan_equipment_list(self):
        self.equipment_list = self.list_instrument.list_connected_devices()
        # To be deleted when actual instrument was used.
        # self.equipment_list = ('USB0::0x0B3E::0x1012::XF001773::0::INSTR',
        #                        'ASRL4::INSTR', 'ASRL8::INSTR',
        #                        'USB0::2391::1543::MY53020107::INSTR',
        #                        'USB0::0x2A8D::0x1301::MY53218004::0::INSTR',
        #                        'USB0::0x0699::0x0408::C014709::0::INSTR',
        #                        'USB0::0x0AAD::0x0197::1329.7002k44-320094::0::INSTR')

    # Start Test
    def start_test(self):

        # Check input parameter
        self.check_input_parameter()

        # Set threading class
        self.measurement_thread = QThread()

        global thread_running

        if not thread_running:
            if self.error_count <= 0:
                self.pushButton_StartTest.setEnabled(False)
                self.eff_meas = Eff_Measurement(PS_USED=self.checkBox_ExtSupUsed.isChecked(), PS_ADD=self.PS_address,
                                                PS_VSTART=self.PS_target_Vin, PS_VMAX=self.PS_limit_Vin,
                                                PS_IMAX=self.PS_limit_Cin, ELOAD_ADD=self.ELoad_address, ELOAD_CHANNEL=self.Eload_channel,
                                                ELOAD_START=self.ELoad_ISTART, ELOAD_MAX=self.ELoad_IMAX,
                                                ELOAD_STEP=self.ELoad_ISTEP,
                                                DMM_VIN_XUSED=self.checkBox_ReadIntVol_PSEquip.isChecked(),
                                                DMM_VIN_ADD=self.DMM_VIN_address,
                                                DMM_CIN_XUSED=self.checkBox_ReadIntCur_PSEquip.isChecked(),
                                                DMM_CIN_ADD=self.DMM_CIN_address,
                                                DMM_VOUT_XUSED=self.checkBox_ReadOutVol_ELoad.isChecked(),
                                                DMM_VOUT_ADD=self.DMM_VOUT_address,
                                                DMM_COUT_XUSED=self.checkBox_ReadOutCur_ELoad.isChecked(),
                                                DMM_COUT_ADD=self.DMM_COUT_address, OUTPUT_DIR=self.output_directory,
                                                OUTPUT_NAME=self.output_filename,
                                                EXP_PDF=self.checkBox_ExpToPdf.isChecked())
                # Indicate thread started

                thread_running = True
                msg_box_auto_close("Test started!!")
                self.eff_meas.moveToThread(self.measurement_thread)
                self.measurement_thread.started.connect(self.eff_meas.fixed_vin_meas)
                self.eff_meas.finished.connect(self.show_test_finish)
                self.eff_meas.finished.connect(self.measurement_thread.quit)
                self.eff_meas.finished.connect(self.eff_meas.deleteLater)
                self.eff_meas.progress.connect(self.progress_bar_update)
                self.eff_meas.generate_report.connect(self.show_generating_report)
                self.eff_meas.error.connect(self.show_error)
                self.eff_meas.error.connect(self.measurement_thread.quit)
                self.eff_meas.error.connect(self.eff_meas.deleteLater)
                self.measurement_thread.finished.connect(self.measurement_thread.deleteLater)
                self.measurement_thread.start()

            else:
                msg_box_ok("Please fill up all the required field")

        # If existing thread, kill it/ wait until current process finish
        else:
            msg_box_ok('Please wait the current measurement finish or click "Abort Test"')

    # Stop Test
    def stop_test(self):

        global thread_running
        if not thread_running:
            msg_box_ok("Info:\n"
                       "- Test is not running.\n"
                       "- No test case to abort!")
        else:
            thread_running = False
            # msg_box_ok("Test Stop!")
            self.progressBar.setProperty("value", 0)
            self.pushButton_StartTest.setEnabled(True)

    # Check and verify input parameter
    def check_input_parameter(self):

        self.error_count = 0

        # ----------------------Get Info from External Power Supply (If selected)-------------------------------------
        if self.checkBox_ExtSupUsed.isChecked():
            # Check PS Address Input
            if self.comboBox_PS_Address.currentIndex() > 0:
                self.PS_address = self.comboBox_PS_Address.currentText()
            else:
                self.error_count += 1

            # Check PS VIN input
            if self.lineEdit_PS_VIN.text() == "":
                self.error_count += 1
                self.lineEdit_PS_VIN.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
            else:
                self.lineEdit_PS_VIN.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                self.PS_target_Vin = float(self.lineEdit_PS_VIN.text())

            # Check PS VIN limit input
            if self.lineEdit_PS_VIN_Limit.text() == "":
                self.error_count += 1
                self.lineEdit_PS_VIN_Limit.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
            else:
                self.lineEdit_PS_VIN_Limit.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                self.PS_limit_Vin = float(self.lineEdit_PS_VIN_Limit.text())

            # Check PS IIN limit input
            if self.lineEdit_PS_Cur_Limit.text() == "":
                self.error_count += 1
                self.lineEdit_PS_Cur_Limit.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
            else:
                self.lineEdit_PS_Cur_Limit.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                self.PS_limit_Cin = float(self.lineEdit_PS_Cur_Limit.text())

        # ---------------------- Get Info from Electronic Load ------------------------------------------------------
        # Check Eload Address Input
        if self.comboBox_ELoad_Address.currentIndex() > 0:
            self.ELoad_address = self.comboBox_ELoad_Address.currentText()
        else:
            self.error_count += 1

        if self.comboBox_ELoad_CH.currentIndex() > 0:
            self.Eload_channel = int(self.comboBox_ELoad_CH.currentText())
        else:
            self.error_count += 1

        # Check Eload parameter
        if self.lineEdit_IMAX.text() == "":
            self.error_count += 1
            self.lineEdit_IMAX.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
        else:
            self.lineEdit_IMAX.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.ELoad_IMAX = float(self.lineEdit_IMAX.text())

        if self.lineEdit_ISTEP.text() == "":
            self.error_count += 1
            self.lineEdit_ISTEP.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
        else:
            self.lineEdit_ISTEP.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.ELoad_ISTEP = float(self.lineEdit_ISTEP.text())

        if self.lineEdit_ISTART.text() == "":
            self.error_count += 1
            self.lineEdit_ISTART.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
        else:
            self.lineEdit_ISTART.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.ELoad_ISTART = float(self.lineEdit_ISTART.text())

        # ---------------------- Get Info from DMMs -------------------------------------
        # DMM1
        if not self.checkBox_ReadIntVol_PSEquip.isChecked():
            if self.comboBox_DMM_VI.currentIndex() > 0:
                self.DMM_VIN_address = self.comboBox_DMM_VI.currentText()
            else:
                self.error_count += 1

        # DMM2
        if not self.checkBox_ReadIntCur_PSEquip.isChecked():
            if self.comboBox_DMM_CI.currentIndex() > 0:
                self.DMM_CIN_address = self.comboBox_DMM_CI.currentText()
            else:
                self.error_count += 1

        # DMM3
        if not self.checkBox_ReadOutVol_ELoad.isChecked():
            if self.comboBox_DMM_VO.currentIndex() > 0:
                self.DMM_VOUT_address = self.comboBox_DMM_VO.currentText()
            else:
                self.error_count += 1

        # DMM4
        if not self.checkBox_ReadOutCur_ELoad.isChecked():
            if self.comboBox_DMM_CO.currentIndex() > 0:
                self.DMM_COUT_address = self.comboBox_DMM_CO.currentText()
            else:
                self.error_count += 1

        # ---------------------- Get info for output directory -------------------------------------
        if self.lineEdit_Out_Directory.text() == "":
            self.error_count += 1
            self.lineEdit_Out_Directory.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
        else:
            self.lineEdit_Out_Directory.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.output_directory = self.lineEdit_Out_Directory.text()

        if self.lineEdit_Out_FileName.text() == "":
            self.error_count += 1
            self.lineEdit_Out_FileName.setStyleSheet("QLineEdit{background-color: rgb(255, 10, 10);}")
        else:
            self.lineEdit_Out_FileName.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.output_filename = self.lineEdit_Out_FileName.text()

    # Report Progress
    def progress_bar_update(self, n):
        self.progressBar.setProperty("value", int(n))

    # Report Error
    def show_error(self, error_number, msg):
        msg_box_ok(f"ERROR: Stopped due to the following error!\n"
                   f"- {msg}")
        self.pushButton_StartTest.setEnabled(True)

    # Report test finish
    def show_test_finish(self):
        global thread_running
        thread_running = False
        msg_box_ok(f"Test completed! Please view the generated report at {self.lineEdit_Out_Directory.text()}.")
        self.pushButton_StartTest.setEnabled(True)

    # Report generating report
    def show_generating_report(self):
        msg_box_auto_close(f"Generating Report")

    # Browse directory
    def browse_directory(self):
        root = tk.Tk()
        root.withdraw()
        target_path = filedialog.askdirectory()
        if target_path:
            self.lineEdit_Out_Directory.setText(target_path)


class Eff_Measurement(QObject):

    # Child Thread
    finished = pyqtSignal()
    generate_report = pyqtSignal()
    progress = pyqtSignal(int)
    error = pyqtSignal(int, str)

    def __init__(self, PS_USED=False, PS_ADD="", PS_VSTART=0, PS_VMAX=0, PS_VSTEP=0,PS_IMAX=0,
                 ELOAD_ADD="", ELOAD_CHANNEL=1, ELOAD_START=0, ELOAD_MAX=0, ELOAD_STEP=0,
                 DMM_VIN_XUSED=False, DMM_VIN_ADD="",
                 DMM_CIN_XUSED=False, DMM_CIN_ADD="",
                 DMM_VOUT_XUSED=False, DMM_VOUT_ADD="",
                 DMM_COUT_XUSED=False, DMM_COUT_ADD="",
                 OUTPUT_DIR="", OUTPUT_NAME="", EXP_PDF=True):

        super().__init__()
        self.ps_used = PS_USED
        self.ps_add = PS_ADD
        self.ps_vstart = PS_VSTART
        self.ps_vmax = PS_VMAX
        self.ps_vstep = PS_VSTEP
        self.ps_imax = PS_IMAX
        self.eload_add = ELOAD_ADD
        self.eload_channel = ELOAD_CHANNEL
        self.eload_istart = ELOAD_START
        self.eload_imax = ELOAD_MAX
        self.eload_istep = ELOAD_STEP
        self.dmm_vin_xused = DMM_VIN_XUSED
        self.dmm_vin_add = DMM_VIN_ADD
        self.dmm_cin_xused = DMM_CIN_XUSED
        self.dmm_cin_add = DMM_CIN_ADD
        self.dmm_vout_xused = DMM_VOUT_XUSED
        self.dmm_vout_add = DMM_VOUT_ADD
        self.dmm_cout_xused = DMM_COUT_XUSED
        self.dmm_cout_add = DMM_COUT_ADD
        self.output_dir = OUTPUT_DIR
        self.output_name = OUTPUT_NAME
        self.exp_pdf = EXP_PDF

        self.vin_measured = []
        self.iin_measured = []
        self.vout_measured = []
        self.iout_measured = []
        self.pin_calculated = []
        self.pout_calculated = []
        self.eff_calculated = []

        self.ext_power_supply = PS_Kikusui_PyVisa.Kikusui_features()  # Kikusui External Power Supply
        self.eload_command = Eload_Chroma_PyVisa.ELOAD_Chroma_features()    # Chroma ELOAD
        self.dmm_vin = DMM_Keysight_PyVisa.Keysight_DMM()
        self.dmm_iin = DMM_Keysight_PyVisa.Keysight_DMM()
        self.dmm_vout = DMM_Keysight_PyVisa.Keysight_DMM()
        self.dmm_iout = DMM_Keysight_PyVisa.Keysight_DMM()

    def fixed_vin_meas(self):
        global thread_running
        error = 0
        error_msg = ""

        # connect ELOAD remotely
        self.eload_command.connect_equipment(target_resource_instr=self.eload_add)
        self.eload_command.config_remote("ON")

        # Count the required loops required
        self.eload_steps = (self.eload_imax - self.eload_istart) / self.eload_istep
        self.eload_steps_round = math.floor(self.eload_steps)

        if self.eload_steps > self.eload_steps_round:
            self.eload_steps_round += 1

        # Turn on external power supply if required
        if self.ps_used:
            self.ext_power_supply.connect_equipment(self.ps_add)
            self.ext_power_supply.set_cv_output(mode='CV', polar='UNIP', out_vol=self.ps_vstart, out_vol_lim=self.ps_vmax, out_cur_lim=self.ps_imax)
            self.status = self.ext_power_supply.on_off_equipment(1)
            time.sleep(1)

            if self.status:
                print(f"Turn on external power supply {self.ps_add}")
            else:
                print(f"Error - function UI-> Eff_measurement -> fixed vin_meas: \n\n"
                      f"Failed to turn on external power supply {self.ps_add}")
                error += 1
                error_msg = f"Failed to turn on external power supply {self.ps_add}"

        # Start looping test
        if error == 0:
            for i in range(self.eload_steps_round + 1):
                if thread_running:                                                    # Check if meas thread is running

                    # Configure ELoad current
                    if self.eload_steps_round > self.eload_steps:
                        eload_current = self.eload_imax
                    else:
                        eload_current = self.eload_istart + (self.eload_istep * i)
                    eload_set_status = self.eload_command.static_load(self.eload_channel, eload_current, "ON")

                    # Calculate measurement progress
                    meas_progress = int((i / self.eload_steps_round) * 100)
                    self.progress.emit(meas_progress)

                    # If error on ELOAD
                    if not eload_set_status:
                        error += 1
                        error_msg = "Unable to configure ELOAD"
                        break

                    time.sleep(3)                                                       # For load current to stabilize

                    # Configure VIN measuring devices
                    if not self.dmm_vin_xused:
                        self.dmm_vin.connect_equipment(self.dmm_vin_add)
                        self.vin_measured.append(self.dmm_vin.meas_vdc()[0])
                    else:
                        self.vin_measured.append(self.ext_power_supply.read_output_supply()[0][0])

                    # Configure IIN measuring devices
                    if not self.dmm_cin_xused:
                        self.dmm_iin.connect_equipment(self.dmm_cin_add)
                        self.iin_measured.append(self.dmm_iin.meas_idc()[0])
                    else:
                        self.iin_measured.append(self.ext_power_supply.read_output_supply()[1][0])

                    # Configure VOUT measuring devices
                    if not self.dmm_vout_xused:
                        self.dmm_vout.connect_equipment(self.dmm_vout_add)
                        self.vout_measured.append(self.dmm_vout.meas_vdc()[0])
                    else:
                        self.vout_measured.append(self.eload_command.voltage_read()[0])

                    # Configure IOUT measuring devices
                    if not self.dmm_cout_xused:
                        self.dmm_iout.connect_equipment(self.dmm_cout_add)
                        self.iout_measured.append(self.dmm_iout.meas_idc()[0])
                    else:
                        self.iout_measured.append(self.eload_command.current_read()[0])

                else:
                    error += 1
                    error_msg = "Test Aborted!"
                    print("Thread Stop")
                    break

        if error:
            self.error.emit(error, error_msg)
        else:
            self.pin_calculated, self.pout_calculated, self.eff_calculated = self.eff_calculation(self.vin_measured, self.iin_measured, self.vout_measured, self.iout_measured)
            self.generate_report.emit()
            error = self.export_result_to_excel(self.vin_measured, self.iin_measured, self.vout_measured, self.iout_measured, self.pin_calculated, self.pout_calculated, self.eff_calculated, self.output_dir, self.output_name, self.exp_pdf)
            if error:
                self.error.emit(error, "Failed to generate report")
            else:
                self.finished.emit()

        self.eload_command.config_remote("OFF")
        self.eload_command.static_load(1, 0, "OFF")
        self.ext_power_supply.on_off_equipment(0)

    def eff_calculation(self, meas_vin, meas_iin, meas_vout, meas_iout):
        InPwrCal = []
        OutPwrCal = []
        EffCal = []
        for i in range(self.eload_steps_round + 1):
            input_power = meas_vin[i]*meas_iin[i]
            output_power = meas_vout[i]*meas_iout[i]
            efficiency = output_power/input_power * 100
            InPwrCal.append("{:.3f}".format(input_power))
            OutPwrCal.append("{:.3f}".format(output_power))
            EffCal.append("{:.3f}".format(efficiency))

        return InPwrCal, OutPwrCal, EffCal

    def export_result_to_excel(self, meas_vin, meas_iin, meas_vout, meas_iout,
                               pin_calculated, pout_calculated, eff_calculated, file_dir, file_name, exp_to_pdf):

        destination_path = f"{file_dir}/"
        destination_excel_file = f"{destination_path}{file_name}.xlsx"

        table_start_row = 7
        table_end_row = len(meas_iout) + table_start_row - 1

        try:
            workbook = Workbook()
            ws = workbook.active

            # Setting up table design parameter
            bold20Calibri = Font(size=20, italic=False, bold=True, name='Calibri')
            normal11Calibri = Font(size=11, italic=False, bold=False, name='Calibri')
            bold11Calibri = Font(size=11, italic=False, bold=True, name='Calibri')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.title = "PWR_EFF_TEST"

            ws.cell(row=1, column=1, value="POWER SUPPLY EFFICIENCY MEASUREMENT (%) ACROSS OUTPUT LOAD CURRENT (A)").font = bold20Calibri
            ws.cell(row=3, column=1, value="TEST DATE:").font = normal11Calibri
            ws.cell(row=3, column=2, value=f"{datetime.date.today()}").font = normal11Calibri
            ws.cell(row=4, column=1, value="TEST NAME:").font = normal11Calibri
            ws.cell(row=4, column=2, value=f"{file_name}").font = normal11Calibri
            ws.cell(row=6, column=1, value='INPUT VOLTAGE (V)').font = bold11Calibri
            ws.cell(row=6, column=1).border = thin_border
            ws.cell(row=6, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=2, value='INPUT CURRENT (A)').font = bold11Calibri
            ws.cell(row=6, column=2).border = thin_border
            ws.cell(row=6, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=3, value='INPUT POWER (W)').font = bold11Calibri
            ws.cell(row=6, column=3).border = thin_border
            ws.cell(row=6, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=4, value='OUTPUT VOLTAGE (V)').font = bold11Calibri
            ws.cell(row=6, column=4).border = thin_border
            ws.cell(row=6, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=5, value='OUTPUT CURRENT (A)').font = bold11Calibri
            ws.cell(row=6, column=5).border = thin_border
            ws.cell(row=6, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=6, value='OUTPUT POWER (W)').font = bold11Calibri
            ws.cell(row=6, column=6).border = thin_border
            ws.cell(row=6, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=7, value='EFFICIENCY (%)').font = bold11Calibri
            ws.cell(row=6, column=7).border = thin_border
            ws.cell(row=6, column=7).alignment = Alignment(horizontal='center')

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20

            # Insert values to table
            for row in range(len(meas_iout)):
                ws.cell(row=row+table_start_row, column=1, value=float(meas_vin[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=2, value=float(meas_iin[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=3, value=float(pin_calculated[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=4, value=float(meas_vout[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=5, value=float(meas_iout[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=6, value=float(pout_calculated[row])).font = normal11Calibri
                ws.cell(row=row+table_start_row, column=7, value=float(eff_calculated[row])).font = normal11Calibri

                for column in range(1, 8):
                    ws.cell(row=row + table_start_row, column=column).border = thin_border

            # Create Chart
            chart = ScatterChart()
            chart.title = "POWER SUPPLY EFFICIENCY MEASUREMENT (%) ACROSS OUTPUT CURRENT (A)"
            chart.style = 13
            chart.x_axis.title = "Output Current (A)"
            chart.x_axis.scaling.min = meas_iout[0]
            chart.x_axis.scaling.max = meas_iout[len(meas_iout)-1]
            chart.y_axis.title = "Efficiency (%)"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100

            xvalues = Reference(ws, min_col=5, min_row=table_start_row, max_row=table_end_row)
            yvalues = Reference(ws, min_col=7, min_row=table_start_row, max_row=table_end_row)
            series = Series(yvalues, xvalues)
            chart.series.append(series)
            chart.height = 13
            chart.width = 25
            ws.add_chart(chart, f"A{table_end_row + 3}")
            workbook.save(destination_excel_file)
            print("Excel Report generated")
            error = False

        except:
            error = True
            print("Excel Report generation failed!!")

        if exp_to_pdf and not error:
            destination_pdf_file = f"{destination_path}{file_name}.pdf"
            print_area = f'A1:G{table_end_row + 29}'
            error = self.export_to_pdf(destination_excel_file, destination_pdf_file, print_area)
            if error:
                print("PDF report generation failed!")
                return error
            else:
                print("PDF report generation succesful!")
        return error

    def export_to_pdf(self, source_path, destination_path, print_area):
        o = win32com.client.Dispatch("Excel.Application")
        o.Visible = False
        o.DisplayAlerts = False
        source_path = r'{}'.format(source_path.replace("/", "\\"))
        destination_path_2 = r'{}'.format(destination_path.replace("/", "\\"))

        try:
            ws_index_list = [1]
            wb = o.Workbooks.Open(source_path)
            time.sleep(3)                       # Delay for workbook to open

            for index in ws_index_list:
                ws = wb.Worksheets[index-1]
                ws.PageSetup.Zoom = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.PrintArea = print_area
            wb.Worksheets(ws_index_list).Select()

            try:
                wb.ActiveSheet.ExportAsFixedFormat(0, destination_path_2)
            except:                                         # To overwrite the same file name
                os.remove(destination_path)
                wb.ActiveSheet.ExportAsFixedFormat(0, destination_path_2)

            print("export to pdf succesfully")
            error = False
            wb.Close(SaveChanges=False)
        except:
            error = True
            print("Failed to generate pdf")
        finally:
            o.Workbooks.Close()

        return error


class ABOUT_UI(QDialog, Support.Ui_Dialog_support):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.label_email.setText("<a href='mailto:chenhui_k029@hotmail.com?"
                                 "subject=Support for Automated SMPS Efficiency Test&"
                                 "body=Hi support, \nI would like to seek for help on ..\n'"
                                 ">Click here to sent an email for support regarding the control interface.")
        self.label_email.setOpenExternalLinks(True)


class MEAS_SETUP_GUIDE(QDialog, Measurement_Setup.Ui_Dialog_Measurement_Setup):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


if __name__ == "__main__":
    meas = Eff_Measurement()
    file_dir = "D:\OneDrive - Continental AG"
    file_name = "test_1"
    meas_vin = [1, 2, 3, 4, 5, 6, 7]
    meas_vout = [1, 2, 3, 4, 5, 6, 7]
    meas_iin = [1, 2, 3, 4, 5, 6, 7]
    meas_iout = [1, 2, 3, 4, 5, 6, 7]
    pin_calculated = [2, 3, 4, 5, 6, 7, 8]
    pout_calculated = [2, 3, 4, 5, 6, 7, 8]
    eff_calculated = [11, 12, 13, 14, 15, 16, 21]
    exp_to_pdf = True

    meas.export_result_to_excel(meas_vin, meas_iin, meas_vout, meas_iout,
                               pin_calculated, pout_calculated, eff_calculated, file_dir, file_name, exp_to_pdf)



























